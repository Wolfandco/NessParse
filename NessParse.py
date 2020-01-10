import openpyxl

## Below is the dictionary for Wolf Comment column (add more as necessary/be careful to not duplicate keys)

wolfComment = {'PMM': 'Patch Management - Missing Patch(es)',
               'PMR': 'Patch Management - Registry Edit(s) Necessary',
               'PMG': 'Patch Management - GPO Change Necessary',
               'AMU': 'Asset Management - Unsupported Software',
               'POR': 'Pass on Reporting',
               'PORI': 'Pass on Reporting - Informational',
               'PORIC': 'Pass on Reporting - Internal Crypto',
               'GHS': 'General Hardening - SMB Signing Not Enforced',
               'GHW': 'General Hardening - Disable LLMNR/NBNS',
               'GHSMB': 'General Hardening - Disable SMBv1',
               'GHC': 'General Hardening - Disable Cisco Smart Install',
               'GHR': 'General Hardening - Registry Change Necessary',
               'GHRB': 'General Hardening - Reboot Necessary',
               'GHM': 'General Hardening - Misconfiguration',
               'GHAV': 'General Hardening - A/V Not Updated And/Or Functioning',
               'GHU': 'General Hardening - Unquoted Service Path(s)'
               }

nessusExpressions = {'AS': 'Apply Security',
                     'AC': 'Apply Cumulative'}

## Instead of hard coding the file below, I need to have it prompt for the file location from the user

wb = openpyxl.load_workbook('FULL PATH TO XLS FILE')
sheet = wb.get_sheet_by_name('SHEET NAME HERE')

print('Reading Table...')

## 

tuple(sheet['A2':'O7305'])

columnOfCellObjects = sheet['A2':'O7305']

for sheet.value in columnOfCellObjects:
    solutionName = sheet.cell(row=sheet['A2':'A7305'], column=11).value
    wcColumn = sheet.cell(row=columnOfCellObjects, column=12).value
    if solutionName in nessusExpressions:
        cellObj.value = wolfComment[PMM]
            